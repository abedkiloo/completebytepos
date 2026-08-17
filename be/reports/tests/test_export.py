"""Coverage-focused tests for report PDF / Excel / CSV export."""

from datetime import date, datetime, timezone as dt_timezone
from decimal import Decimal
from io import BytesIO
from unittest.mock import Mock

from django.test import SimpleTestCase
from django.utils import timezone
from openpyxl import load_workbook
from rest_framework import status

from reports.export import (
    UnsupportedExportFormat,
    export_filename,
    flatten_payload,
    humanize_key,
    is_scalar,
    maybe_export_report,
    normalize_export_format,
    render_csv,
    render_pdf,
    render_report_bytes,
    render_xlsx,
    report_file_response,
    scalarize,
    sheet_title,
)


SAMPLE_SALES = {
    'summary': {
        'total_sales': 2,
        'total_revenue': Decimal('350.50'),
        'closed': True,
        'nested': {'tickets': 2, 'voids': 0},
    },
    'by_payment_method': [
        {'payment_method': 'cash', 'count': 1, 'total': 200},
        {'payment_method': 'mpesa', 'count': 1, 'total': Decimal('150.50')},
    ],
    'daily_breakdown': [
        {'day': date(2026, 8, 17), 'count': 2, 'total': 350.5},
    ],
    'notes': ['late entry', 'backfill'],
    'empty_list': [],
    'generated_at': datetime(2026, 8, 17, 12, 0, tzinfo=dt_timezone.utc),
}


class NormalizeFormatTests(SimpleTestCase):
    def test_blank_is_none(self):
        self.assertIsNone(normalize_export_format(None))
        self.assertIsNone(normalize_export_format(''))
        self.assertIsNone(normalize_export_format('  '))

    def test_aliases(self):
        self.assertEqual(normalize_export_format('PDF'), 'pdf')
        self.assertEqual(normalize_export_format('excel'), 'xlsx')
        self.assertEqual(normalize_export_format('XLS'), 'xlsx')
        self.assertEqual(normalize_export_format('csv'), 'csv')

    def test_unsupported(self):
        with self.assertRaises(UnsupportedExportFormat):
            normalize_export_format('pptx')


class ScalarizeTests(SimpleTestCase):
    def test_primitives(self):
        self.assertTrue(is_scalar(None))
        self.assertEqual(scalarize(None), '')
        self.assertEqual(scalarize(True), 'Yes')
        self.assertEqual(scalarize(False), 'No')
        self.assertEqual(scalarize(Decimal('1.5')), 1.5)
        self.assertEqual(scalarize('cash'), 'cash')
        self.assertEqual(scalarize(3), 3)

    def test_dates(self):
        self.assertEqual(scalarize(date(2026, 1, 2)), '2026-01-02')
        text = scalarize(datetime(2026, 1, 2, 15, 4, tzinfo=dt_timezone.utc))
        self.assertIn('2026-01-02', text)

    def test_collections(self):
        self.assertEqual(scalarize([]), '')
        self.assertEqual(scalarize(['a', 'b']), 'a, b')
        self.assertEqual(scalarize([{'id': 1}]), '1 rows')
        self.assertIn('a=1', scalarize({'a': 1, 'b': 2}))
        long_dict = {f'k{i}': i for i in range(12)}
        self.assertEqual(scalarize(long_dict).count('='), 8)
        fallback = scalarize(object())
        self.assertTrue(isinstance(fallback, str) and fallback)

    def test_naive_datetime(self):
        text = scalarize(datetime(2026, 3, 4, 8, 9))
        self.assertIn('2026-03-04', text)


class HumanizeTests(SimpleTestCase):
    def test_humanize_and_sheet_title(self):
        self.assertEqual(humanize_key('by_payment_method'), 'By Payment Method')
        self.assertEqual(humanize_key(''), 'Value')
        self.assertEqual(sheet_title('summary'), 'Summary')
        long_name = 'x' * 80
        self.assertEqual(len(sheet_title(long_name)), 31)
        self.assertEqual(sheet_title('bad:name?'), 'Bad Name')
        self.assertEqual(sheet_title('???'), 'Sheet')

    def test_filename(self):
        when = timezone.make_aware(datetime(2026, 8, 17, 9, 0))
        self.assertEqual(
            export_filename('sales/overview', 'xlsx', when=when),
            'sales_overview_2026-08-17.xlsx',
        )
        self.assertEqual(export_filename('', 'pdf', when=when), 'report_2026-08-17.pdf')


class FlattenPayloadTests(SimpleTestCase):
    def test_none_and_scalar(self):
        none_sheets = flatten_payload(None)
        self.assertEqual(none_sheets[0][0], 'summary')
        scalar_sheets = flatten_payload('hello')
        self.assertEqual(scalar_sheets[0][2][0][1], 'hello')

    def test_list_payload(self):
        sheets = flatten_payload([{'sku': 'A'}, 'plain'])
        self.assertEqual(sheets[0][0], 'rows')
        self.assertGreaterEqual(len(sheets[0][2]), 2)

    def test_nested_sales_shape(self):
        sheets = {name: (headers, rows) for name, headers, rows in flatten_payload(SAMPLE_SALES)}
        self.assertIn('summary', sheets)
        self.assertIn('by_payment_method', sheets)
        self.assertIn('daily_breakdown', sheets)
        self.assertIn('notes', sheets)
        self.assertTrue(any('Total Revenue' in row for row in sheets['summary'][1]))

    def test_other_nested_values(self):
        payload = {
            'meta': {
                'weird': [{'x': 1}, {'x': 2}],
                'blob': {'a': [1, 2]},
                'empty': {},
            }
        }
        sheets = flatten_payload(payload)
        names = [name for name, _, _ in sheets]
        self.assertTrue(any('weird' in name for name in names))
        summary = next(rows for name, _, rows in sheets if name == 'summary')
        summary_text = ' '.join(str(cell) for row in summary for cell in row)
        self.assertIn('Blob', summary_text)

    def test_empty_dict_is_placeholder(self):
        sheets = flatten_payload({})
        self.assertEqual(sheets[0][0], 'summary')
        self.assertEqual(sheets[0][2][0][1], 'Empty')


class RenderTests(SimpleTestCase):
    def test_csv_contains_headers_and_rows(self):
        body = render_csv(SAMPLE_SALES).decode('utf-8-sig')
        self.assertIn('By Payment Method', body)
        self.assertIn('cash', body)
        self.assertIn('mpesa', body)

    def test_xlsx_has_named_sheets(self):
        content = render_xlsx(SAMPLE_SALES, title='Sales Report')
        workbook = load_workbook(BytesIO(content))
        self.assertIn('Summary', workbook.sheetnames)
        self.assertTrue(any('Payment' in name for name in workbook.sheetnames))
        self.assertEqual(workbook.active['A1'].value, 'Sales Report')

    def test_xlsx_unique_sheet_names(self):
        payload = {
            'items': [{'v': 1}],
            'items_extra': [{'v': 2}],
        }
        content = render_xlsx(payload, title='Dup')
        workbook = load_workbook(BytesIO(content))
        self.assertGreaterEqual(len(set(workbook.sheetnames)), 2)

    def test_xlsx_duplicate_titles_get_suffix(self):
        # Force colliding sheet titles after humanize (slash, underscore, hyphen).
        payload = {
            'foo/bar': [{'v': 1}],
            'foo_bar': [{'v': 2}],
            'foo-bar': [{'v': 3}],
        }
        content = render_xlsx(payload, title='Clash')
        workbook = load_workbook(BytesIO(content))
        self.assertEqual(len(workbook.sheetnames), 3)
        self.assertEqual(len(set(workbook.sheetnames)), 3)

    def test_pdf_magic(self):
        content = render_pdf(SAMPLE_SALES, title='Sales Report')
        self.assertTrue(content.startswith(b'%PDF'))
        self.assertGreater(len(content), 200)

    def test_empty_payload_pdf(self):
        content = render_pdf({})
        self.assertTrue(content.startswith(b'%PDF'))

    def test_wide_table_pdf_truncates_columns(self):
        row = {f'col_{i}': i for i in range(12)}
        content = render_pdf({'rows': [row]}, title='Wide')
        self.assertTrue(content.startswith(b'%PDF'))

    def test_render_report_bytes_routes(self):
        self.assertTrue(render_report_bytes(SAMPLE_SALES, 'pdf').startswith(b'%PDF'))
        self.assertGreater(len(render_report_bytes(SAMPLE_SALES, 'xlsx')), 100)
        self.assertIn(b'cash', render_report_bytes(SAMPLE_SALES, 'csv'))
        with self.assertRaises(UnsupportedExportFormat):
            render_report_bytes(SAMPLE_SALES, 'gif')


class ResponseTests(SimpleTestCase):
    def test_file_response_headers(self):
        response = report_file_response(
            SAMPLE_SALES, fmt='pdf', title='Sales', slug='sales'
        )
        self.assertEqual(response['Content-Type'], 'application/pdf')
        self.assertIn('attachment', response['Content-Disposition'])
        self.assertEqual(response['X-Report-Export'], 'pdf')

    def test_csv_body_override(self):
        response = report_file_response(
            SAMPLE_SALES,
            fmt='csv',
            title='Sales',
            slug='sales',
            csv_body='Sales staff performance report\nstaff_b\n',
        )
        self.assertIn('Sales staff performance report', response.content.decode('utf-8-sig'))

    def test_csv_body_bytes_override(self):
        response = report_file_response(
            SAMPLE_SALES,
            fmt='csv',
            title='Sales',
            slug='sales',
            csv_body=b'raw,csv',
        )
        self.assertIn(b'raw,csv', response.content)

    def test_maybe_export_none_without_format(self):
        request = Mock()
        request.query_params = {}
        request.GET = {}
        self.assertIsNone(maybe_export_report(request, SAMPLE_SALES, title='T', slug='sales'))

    def test_maybe_export_pdf(self):
        request = Mock()
        request.query_params = {'format': 'pdf'}
        request.GET = {}
        response = maybe_export_report(request, SAMPLE_SALES, title='Sales', slug='sales')
        self.assertEqual(response['Content-Type'], 'application/pdf')

    def test_maybe_export_excel_alias(self):
        request = Mock()
        request.query_params = {'format': 'excel'}
        request.GET = {}
        response = maybe_export_report(request, SAMPLE_SALES, title='Sales', slug='sales')
        self.assertIn('spreadsheet', response['Content-Type'])

    def test_maybe_export_invalid(self):
        request = Mock()
        request.query_params = {'format': 'ppt'}
        request.GET = {}
        response = maybe_export_report(request, SAMPLE_SALES, title='Sales', slug='sales')
        self.assertEqual(response.status_code, status.HTTP_400_BAD_REQUEST)
        self.assertIn('error', response.data)

    def test_maybe_export_without_request(self):
        self.assertIsNone(maybe_export_report(None, SAMPLE_SALES, title='T', slug='s'))

    def test_maybe_export_uses_get_fallback(self):
        request = Mock()
        request.query_params = {}
        request.GET = {'format': 'csv'}
        response = maybe_export_report(request, SAMPLE_SALES, title='Sales', slug='sales')
        self.assertIn('text/csv', response['Content-Type'])

    def test_maybe_export_csv_body_only_for_csv(self):
        request = Mock()
        request.query_params = {'format': 'pdf'}
        request.GET = {}
        response = maybe_export_report(
            request,
            SAMPLE_SALES,
            title='Sales',
            slug='sales',
            csv_body='should-not-appear',
        )
        self.assertTrue(response.content.startswith(b'%PDF'))
        self.assertNotIn(b'should-not-appear', response.content)
