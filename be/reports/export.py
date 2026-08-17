"""Export report JSON payloads as PDF, Excel, or CSV."""

from __future__ import annotations

import csv
import io
import re
from datetime import date, datetime
from decimal import Decimal
from typing import Any, Iterable, Optional

from django.http import HttpResponse
from django.utils import timezone
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER, TA_LEFT
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import inch
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

CONTENT_TYPES = {
    'pdf': 'application/pdf',
    'xlsx': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    'csv': 'text/csv; charset=utf-8',
}

_FORMAT_ALIASES = {
    'excel': 'xlsx',
    'xls': 'xlsx',
    'xlsx': 'xlsx',
    'pdf': 'pdf',
    'csv': 'csv',
}


class UnsupportedExportFormat(ValueError):
    """Raised when format= is present but not a supported export type."""


def normalize_export_format(raw: Any) -> Optional[str]:
    if raw is None:
        return None
    text = str(raw).strip().lower()
    if not text:
        return None
    fmt = _FORMAT_ALIASES.get(text)
    if not fmt:
        raise UnsupportedExportFormat(
            f'Unsupported export format "{raw}". Use pdf, xlsx, or csv.'
        )
    return fmt


def is_scalar(value: Any) -> bool:
    return value is None or isinstance(
        value, (str, int, float, bool, Decimal, datetime, date)
    )


def scalarize(value: Any) -> Any:
    if value is None:
        return ''
    if isinstance(value, bool):
        return 'Yes' if value else 'No'
    if isinstance(value, Decimal):
        return float(value)
    if isinstance(value, datetime):
        if timezone.is_aware(value):
            value = timezone.localtime(value)
        return value.strftime('%Y-%m-%d %H:%M')
    if isinstance(value, date):
        return value.isoformat()
    if is_scalar(value):
        return value
    if isinstance(value, (list, tuple)):
        if not value:
            return ''
        if all(is_scalar(item) for item in value):
            return ', '.join(str(scalarize(item)) for item in value)
        return f'{len(value)} rows'
    if isinstance(value, dict):
        return ', '.join(f'{k}={scalarize(v)}' for k, v in list(value.items())[:8])
    return str(value)


def humanize_key(key: Any) -> str:
    text = re.sub(r'[_-]+', ' ', str(key or '').strip())
    text = re.sub(r'\s+', ' ', text)
    return text.title() if text else 'Value'


def sheet_title(name: str) -> str:
    title = re.sub(r'[\\/*?:\[\]]', ' ', humanize_key(name)).strip() or 'Sheet'
    return title[:31]


def export_filename(slug: str, fmt: str, when: Optional[datetime] = None) -> str:
    stamp = timezone.localtime(when or timezone.now()).strftime('%Y-%m-%d')
    safe = re.sub(r'[^\w.-]+', '_', slug or 'report').strip('_') or 'report'
    ext = 'xlsx' if fmt == 'xlsx' else fmt
    return f'{safe}_{stamp}.{ext}'


def _table_from_records(records: Iterable[dict]) -> tuple[list[str], list[list[Any]]]:
    rows = [row for row in records if isinstance(row, dict)]
    if not rows:
        return ['Value'], [['No rows']]
    keys: list[str] = []
    seen = set()
    for row in rows:
        for key in row.keys():
            if key not in seen:
                seen.add(key)
                keys.append(str(key))
    body = [[scalarize(row.get(key)) for key in keys] for row in rows]
    headers = [humanize_key(key) for key in keys]
    return headers, body


def flatten_payload(payload: Any) -> list[tuple[str, list[str], list[list[Any]]]]:
    """
    Turn a report dict into named tables: (sheet_name, headers, rows).
    """
    if payload is None:
        payload = {}
    if isinstance(payload, list):
        headers, body = _table_from_records(
            [item if isinstance(item, dict) else {'value': item} for item in payload]
        )
        return [('rows', headers, body)]
    if not isinstance(payload, dict):
        return [('summary', ['Metric', 'Value'], [['Report', scalarize(payload)]])]

    sheets: list[tuple[str, list[str], list[list[Any]]]] = []
    summary_rows: list[list[Any]] = []

    def add_summary(section: str, metric: str, value: Any) -> None:
        summary_rows.append([humanize_key(section), humanize_key(metric), scalarize(value)])

    for key, value in payload.items():
        if isinstance(value, dict):
            nested_lists = []
            nested_scalars = []
            nested_other = []
            for nested_key, nested_value in value.items():
                if isinstance(nested_value, list):
                    nested_lists.append((nested_key, nested_value))
                elif is_scalar(nested_value) or (
                    isinstance(nested_value, dict)
                    and nested_value
                    and all(is_scalar(v) for v in nested_value.values())
                ):
                    nested_scalars.append((nested_key, nested_value))
                else:
                    nested_other.append((nested_key, nested_value))
            for nested_key, nested_value in nested_scalars:
                if isinstance(nested_value, dict):
                    for inner_key, inner_value in nested_value.items():
                        add_summary(key, f'{nested_key} {inner_key}', inner_value)
                else:
                    add_summary(key, nested_key, nested_value)
            for nested_key, nested_value in nested_lists:
                records = [
                    item if isinstance(item, dict) else {'value': item}
                    for item in nested_value
                ]
                headers, body = _table_from_records(records)
                sheets.append((f'{key}_{nested_key}', headers, body))
            for nested_key, nested_value in nested_other:
                add_summary(key, nested_key, nested_value)
        elif isinstance(value, list):
            records = [
                item if isinstance(item, dict) else {'value': item} for item in value
            ]
            headers, body = _table_from_records(records)
            sheets.append((str(key), headers, body))
        else:
            add_summary('summary', key, value)

    if summary_rows:
        sheets.insert(0, ('summary', ['Section', 'Metric', 'Value'], summary_rows))
    if not sheets:
        sheets.append(('summary', ['Metric', 'Value'], [['Report', 'Empty']]))
    return sheets


def render_csv(payload: Any) -> bytes:
    output = io.StringIO()
    writer = csv.writer(output)
    for index, (name, headers, rows) in enumerate(flatten_payload(payload)):
        if index:
            writer.writerow([])
        writer.writerow([humanize_key(name)])
        writer.writerow(headers)
        for row in rows:
            writer.writerow(row)
    return output.getvalue().encode('utf-8-sig')


def render_xlsx(payload: Any, title: str = 'Report') -> bytes:
    workbook = Workbook()
    default = workbook.active
    sheets = flatten_payload(payload)
    used_names: set[str] = set()
    for index, (name, headers, rows) in enumerate(sheets):
        label = sheet_title(name)
        unique = label
        suffix = 1
        while unique.lower() in used_names:
            suffix += 1
            unique = f'{label[:28]}_{suffix}'
        used_names.add(unique.lower())
        worksheet = default if index == 0 else workbook.create_sheet()
        worksheet.title = unique
        worksheet['A1'] = title
        worksheet['A1'].font = Font(bold=True, size=14)
        header_fill = PatternFill('solid', fgColor='1F4E79')
        header_font = Font(bold=True, color='FFFFFF')
        for col, header in enumerate(headers, start=1):
            cell = worksheet.cell(row=3, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')
        for row_index, row in enumerate(rows, start=4):
            for col, value in enumerate(row, start=1):
                worksheet.cell(row=row_index, column=col, value=value)
        for col in range(1, len(headers) + 1):
            worksheet.column_dimensions[get_column_letter(col)].width = 18
    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def render_pdf(payload: Any, title: str = 'Report') -> bytes:
    buffer = io.BytesIO()
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'ReportTitle',
        parent=styles['Heading1'],
        fontSize=16,
        alignment=TA_CENTER,
        spaceAfter=12,
    )
    heading_style = ParagraphStyle(
        'ReportHeading',
        parent=styles['Heading2'],
        fontSize=11,
        alignment=TA_LEFT,
        spaceBefore=10,
        spaceAfter=6,
    )
    doc = SimpleDocTemplate(
        buffer,
        pagesize=landscape(A4),
        leftMargin=0.5 * inch,
        rightMargin=0.5 * inch,
        topMargin=0.5 * inch,
        bottomMargin=0.5 * inch,
    )
    elements = [
        Paragraph(str(title or 'Report'), title_style),
        Paragraph(
            f'Generated {timezone.localtime().strftime("%Y-%m-%d %H:%M")}',
            styles['Normal'],
        ),
        Spacer(1, 0.15 * inch),
    ]
    for name, headers, rows in flatten_payload(payload):
        display_headers = headers[:8]
        display_rows = [row[:8] for row in rows]
        elements.append(Paragraph(humanize_key(name), heading_style))
        table_data = [display_headers] + [
            [str(cell) if cell is not None else '' for cell in row] for row in display_rows
        ]
        col_width = 10 * inch / max(len(display_headers), 1)
        table = Table(table_data, colWidths=[col_width] * len(display_headers), repeatRows=1)
        table.setStyle(
            TableStyle(
                [
                    ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1F4E79')),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, 0), (-1, -1), 7),
                    ('GRID', (0, 0), (-1, -1), 0.25, colors.HexColor('#CCCCCC')),
                    ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                    ('LEFTPADDING', (0, 0), (-1, -1), 4),
                    ('RIGHTPADDING', (0, 0), (-1, -1), 4),
                    ('BACKGROUND', (0, 1), (-1, -1), colors.white),
                    (
                        'ROWBACKGROUNDS',
                        (0, 1),
                        (-1, -1),
                        [colors.white, colors.HexColor('#F4F7FB')],
                    ),
                ]
            )
        )
        elements.append(table)
        elements.append(Spacer(1, 0.12 * inch))
    doc.build(elements)
    return buffer.getvalue()


def render_report_bytes(payload: Any, fmt: str, title: str = 'Report') -> bytes:
    if fmt == 'pdf':
        return render_pdf(payload, title=title)
    if fmt == 'xlsx':
        return render_xlsx(payload, title=title)
    if fmt == 'csv':
        return render_csv(payload)
    raise UnsupportedExportFormat(
        f'Unsupported export format "{fmt}". Use pdf, xlsx, or csv.'
    )


def report_file_response(
    payload: Any,
    *,
    fmt: str,
    title: str,
    slug: str,
    csv_body: Optional[str] = None,
) -> HttpResponse:
    filename = export_filename(slug, fmt)
    if fmt == 'csv' and csv_body is not None:
        content = csv_body.encode('utf-8-sig') if isinstance(csv_body, str) else csv_body
    else:
        content = render_report_bytes(payload, fmt, title=title)
    response = HttpResponse(content, content_type=CONTENT_TYPES[fmt])
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    response['X-Report-Export'] = fmt
    return response


def maybe_export_report(
    request,
    payload: Any,
    *,
    title: str,
    slug: str,
    csv_body: Optional[str] = None,
):
    raw = None
    if request is not None:
        raw = request.query_params.get('format') or request.GET.get('format')
    try:
        fmt = normalize_export_format(raw)
    except UnsupportedExportFormat as exc:
        from rest_framework import status
        from rest_framework.response import Response

        return Response({'error': str(exc)}, status=status.HTTP_400_BAD_REQUEST)
    if not fmt:
        return None
    return report_file_response(
        payload,
        fmt=fmt,
        title=title,
        slug=slug,
        csv_body=csv_body if fmt == 'csv' else None,
    )
