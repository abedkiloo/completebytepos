#!/usr/bin/env python3
"""
Generate CompleteBytePOS Manager & Super Admin onboarding workbook (Excel).

Usage:
  python scripts/generate_user_manual.py
  python scripts/generate_user_manual.py --screenshots-dir docs/user-manual/screenshots

If screenshots exist in docs/user-manual/screenshots/, they are embedded in the workbook.
Run scripts/capture_user_manual_screenshots.js after starting the app to capture real UI shots.
"""
from __future__ import annotations

import argparse
from datetime import date
from pathlib import Path

from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from PIL import Image as PILImage, ImageDraw, ImageFont

ROOT = Path(__file__).resolve().parents[1]
OUT_DIR = ROOT / "docs" / "user-manual"
DEFAULT_SHOTS = OUT_DIR / "screenshots"
OUT_FILE = OUT_DIR / "CompleteBytePOS-Manager-Admin-Manual.xlsx"

HEADER_FILL = PatternFill("solid", fgColor="1E3A5F")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
TITLE_FONT = Font(bold=True, size=16, color="1E3A5F")
SUBTITLE_FONT = Font(bold=True, size=12, color="334155")
WRAP = Alignment(wrap_text=True, vertical="top")
THIN = Side(style="thin", color="CBD5E1")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

COLUMNS = [
    ("section", "Section", 14),
    ("module", "Module", 16),
    ("route", "Page / URL", 22),
    ("task", "Task", 28),
    ("role", "Who", 14),
    ("steps", "Step-by-step instructions", 52),
    ("result", "Expected result", 28),
    ("notes", "Tips & notes", 28),
    ("shot_file", "Screenshot file", 24),
]


def make_placeholder(path: Path, title: str, subtitle: str) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    if path.exists():
        return
    img = PILImage.new("RGB", (960, 540), color=(241, 245, 249))
    draw = ImageDraw.Draw(img)
    draw.rectangle([(24, 24), (936, 516)], outline=(148, 163, 184), width=2)
    draw.text((48, 48), "CompleteByte POS", fill=(30, 58, 95))
    draw.text((48, 100), title, fill=(15, 23, 42))
    draw.text((48, 150), subtitle, fill=(100, 116, 139))
    draw.text(
        (48, 220),
        "Replace: run capture script with app running on :3000",
        fill=(100, 116, 139),
    )
    draw.text((48, 260), f"File: {path.name}", fill=(100, 116, 139))
    img.save(path, "PNG")


MANUAL_ROWS = [
    # --- Getting started ---
    {
        "sheet": "Getting Started",
        "section": "Login",
        "module": "Authentication",
        "route": "/login",
        "task": "Sign in as Manager",
        "role": "Manager",
        "steps": (
            "1. Open the POS URL in Chrome or Edge (e.g. http://your-server:3000).\n"
            "2. Enter username: manager\n"
            "3. Enter password: manager123 (change after first login in production).\n"
            "4. Click Sign in.\n"
            "5. You land on the Dashboard."
        ),
        "result": "Dashboard loads with manager quick actions (POS, Products, Inventory, Reports).",
        "notes": "Sales staff use username sales. Super Admin uses admin.",
        "shot_file": "01-login.png",
    },
    {
        "sheet": "Getting Started",
        "section": "Login",
        "module": "Authentication",
        "route": "/login",
        "task": "Sign in as Super Admin",
        "role": "Super Admin",
        "steps": (
            "1. Open /login.\n"
            "2. Username: admin — Password: admin123.\n"
            "3. Sign in → Dashboard shows admin shortcuts (Users, Module Settings)."
        ),
        "result": "Full navigation including Settings and User Management.",
        "notes": "Only Super Admin can manage users, roles, modules, branches, and system rules.",
        "shot_file": "02-dashboard-admin.png",
    },
    {
        "sheet": "Getting Started",
        "section": "Navigation",
        "module": "Layout",
        "route": "/",
        "task": "Use the sidebar",
        "role": "Manager, Super Admin",
        "steps": (
            "1. Left sidebar groups features: Main, Inventory, Sales, Finance, Reports.\n"
            "2. Click a section heading to expand/collapse.\n"
            "3. Active page is highlighted.\n"
            "4. On mobile, open the menu icon (☰) to show the sidebar.\n"
            "5. Sign out from the profile menu (top-right)."
        ),
        "result": "Reach any module your role is allowed to access.",
        "notes": "If a link is missing, the module may be disabled in Module Settings (Super Admin).",
        "shot_file": "03-sidebar-manager.png",
    },
    {
        "sheet": "Getting Started",
        "section": "Roles",
        "module": "Security",
        "route": "—",
        "task": "Understand the three default roles",
        "role": "All",
        "steps": (
            "Super Admin — full control: users, roles, module install, system settings, branches, financial approvals.\n"
            "Manager — daily operations: catalog, stock, POS, reports, expenses/income entry, product/inventory approvals. Cannot manage users or approve expenses/income by default.\n"
            "Sales Personnel — POS, customers, limited catalog (if allowed). Cannot change prices, costs, or stock."
        ),
        "result": "Assign work to the right person and know when to escalate to Super Admin.",
        "notes": "Custom roles can be created by Super Admin under Roles.",
        "shot_file": "04-roles-reference.png",
    },
    # --- Dashboard ---
    {
        "sheet": "Manager Daily Ops",
        "section": "Dashboard",
        "module": "Dashboard",
        "route": "/",
        "task": "Review morning KPIs",
        "role": "Manager",
        "steps": (
            "1. Open Dashboard after login.\n"
            "2. Check Today's sales, month totals, and profit (if enabled).\n"
            "3. Review low-stock alerts — click through to Inventory or Products.\n"
            "4. Use Quick Actions: Start POS, Terminal POS, Products, Inventory, Reports."
        ),
        "result": "Operational picture before opening the register.",
        "notes": "KPIs respect branch filter when multi-branch is enabled.",
        "shot_file": "05-dashboard-manager.png",
    },
    # --- POS ---
    {
        "sheet": "Sales & POS",
        "section": "Retail POS",
        "module": "Sales",
        "route": "/pos",
        "task": "Complete a cash sale",
        "role": "Manager, Sales",
        "steps": (
            "1. Sidebar → POS (full-screen).\n"
            "2. Search or browse products; click to add to cart.\n"
            "3. For variant products: pick size/color in the popup, set quantity, Add to cart.\n"
            "4. Optional: attach customer, apply discount/tax if module allows.\n"
            "5. Choose payment method (Cash, M-Pesa, etc.).\n"
            "6. Enter amount received → Complete sale.\n"
            "7. Receipt dialog — print or close."
        ),
        "result": "Sale recorded; stock reduced; receipt available.",
        "notes": "Out-of-stock items are blocked when Validate stock is on in Sales module settings.",
        "shot_file": "06-pos-retail.png",
    },
    {
        "sheet": "Sales & POS",
        "section": "Terminal POS",
        "module": "Sales",
        "route": "/pos/billing",
        "task": "Hold and resume a cart",
        "role": "Manager, Sales",
        "steps": (
            "1. Open Terminal POS from sidebar.\n"
            "2. Add items to the cart.\n"
            "3. Cart auto-saves as a holding invoice.\n"
            "4. Leave and return later — cart restores.\n"
            "5. Checkout when ready: payment method, amount, confirm."
        ),
        "result": "Holding sale converts to completed sale without losing items.",
        "notes": "Useful for busy counters and credit customers.",
        "shot_file": "07-terminal-pos.png",
    },
    {
        "sheet": "Sales & POS",
        "section": "Sales history",
        "module": "Sales",
        "route": "/sales",
        "task": "Refund a sale",
        "role": "Manager, Super Admin",
        "steps": (
            "1. Open Sales history.\n"
            "2. Filter by date, cashier, or sale number.\n"
            "3. Open the sale row → View details.\n"
            "4. Click Refund (manager always allowed; sales only with permission).\n"
            "5. Confirm — stock is restored per line."
        ),
        "result": "Sale marked refunded; inventory adjusted.",
        "notes": "Sales staff cannot refund unless granted sales.refund permission.",
        "shot_file": "08-sales-history.png",
    },
    {
        "sheet": "Sales & POS",
        "section": "Normal sale",
        "module": "Sales",
        "route": "/normal-sale",
        "task": "Create an invoiced sale with partial payment",
        "role": "Manager",
        "steps": (
            "1. Open Normal Sale.\n"
            "2. Select customer (required for credit).\n"
            "3. Add products and quantities.\n"
            "4. Set tax/discount if enabled.\n"
            "5. Submit — invoice created with balance due.\n"
            "6. Record payments later under Invoices."
        ),
        "result": "Invoice linked to customer with outstanding balance tracked.",
        "notes": "Requires Invoicing module and normal_sale feature enabled.",
        "shot_file": "09-normal-sale.png",
    },
    # --- Products ---
    {
        "sheet": "Products & Inventory",
        "section": "Products",
        "module": "Products",
        "route": "/products",
        "task": "Add a new product",
        "role": "Manager",
        "steps": (
            "1. Products → Add product.\n"
            "2. Fill name, SKU, category, unit of measure.\n"
            "3. Set MRP, selling price, cost (manager can edit; sales cannot).\n"
            "4. Set opening stock if tracking inventory.\n"
            "5. Upload image (optional).\n"
            "6. Save — if Maker-checker is on, price/stock changes may queue for approval."
        ),
        "result": "Product appears in catalog and POS (after approval if queued).",
        "notes": "Pending changes show a badge on the product list.",
        "shot_file": "10-products-list.png",
    },
    {
        "sheet": "Products & Inventory",
        "section": "Products",
        "module": "Products",
        "route": "/products",
        "task": "Adjust stock from product list",
        "role": "Manager",
        "steps": (
            "1. Products list → find the product.\n"
            "2. Click stock quantity or row menu → Adjust stock.\n"
            "3. Enter +/- quantity and reason.\n"
            "4. Submit — may go to Pending approvals if Maker-checker is on."
        ),
        "result": "Stock updated (or queued) without opening full edit form.",
        "notes": "Stock column is read-only; always use Adjust stock action.",
        "shot_file": "11-stock-adjust-modal.png",
    },
    {
        "sheet": "Products & Inventory",
        "section": "Categories",
        "module": "Products",
        "route": "/categories",
        "task": "Manage categories",
        "role": "Manager",
        "steps": (
            "1. Open Categories.\n"
            "2. Add category or subcategory.\n"
            "3. Set active/inactive.\n"
            "4. Deactivate/delete may require checker approval when Maker-checker is on."
        ),
        "result": "POS and reports group products by category.",
        "notes": "Inactive categories hide products from POS depending on settings.",
        "shot_file": "12-categories.png",
    },
    {
        "sheet": "Products & Inventory",
        "section": "Variants",
        "module": "Products",
        "route": "/product-attributes",
        "task": "Set up sizes and colors",
        "role": "Manager",
        "steps": (
            "1. Open Sizes & Colors (Product attributes).\n"
            "2. Create sizes (S, M, L) and colors.\n"
            "3. In product form, enable variants and assign size/color combinations.\n"
            "4. Set price and stock per variant row."
        ),
        "result": "POS prompts for size/color when adding variant products.",
        "notes": "Requires product_variants feature in Module Settings.",
        "shot_file": "13-product-attributes.png",
    },
    {
        "sheet": "Products & Inventory",
        "section": "Barcodes",
        "module": "Barcodes",
        "route": "/barcodes",
        "task": "Print product labels",
        "role": "Manager",
        "steps": (
            "1. Open Barcodes / Labels.\n"
            "2. Search products or generate missing barcodes.\n"
            "3. Select items → Preview → Print.\n"
            "4. Use barcode scanner at POS to add items by SKU/barcode."
        ),
        "result": "Shelf labels with barcode/QR for scanning.",
        "notes": "Requires barcodes module enabled.",
        "shot_file": "14-barcodes.png",
    },
    # --- Inventory ---
    {
        "sheet": "Products & Inventory",
        "section": "Inventory",
        "module": "Stock",
        "route": "/inventory",
        "task": "Record a stock purchase",
        "role": "Manager",
        "steps": (
            "1. Inventory → Stock purchase (or Purchase tab).\n"
            "2. Select product (and variant if applicable).\n"
            "3. Enter quantity received and unit cost.\n"
            "4. Add supplier reference and notes.\n"
            "5. Save — queues for approval if Maker-checker is on."
        ),
        "result": "Stock increases; weighted average cost updated.",
        "notes": "Use for goods received from suppliers.",
        "shot_file": "15-inventory-purchase.png",
    },
    {
        "sheet": "Products & Inventory",
        "section": "Inventory",
        "module": "Stock",
        "route": "/inventory",
        "task": "Transfer stock between branches",
        "role": "Manager",
        "steps": (
            "1. Inventory → Stock transfer.\n"
            "2. Select product, quantity, source and destination branch.\n"
            "3. Confirm transfer.\n"
            "4. Review movement history for audit trail."
        ),
        "result": "Stock moves between branches with paired movements.",
        "notes": "Requires multi_branch_support feature.",
        "shot_file": "16-inventory-transfer.png",
    },
    {
        "sheet": "Products & Inventory",
        "section": "Inventory",
        "module": "Stock",
        "route": "/inventory",
        "task": "Monitor low and out-of-stock",
        "role": "Manager",
        "steps": (
            "1. Inventory → Low stock tab.\n"
            "2. Review items below threshold.\n"
            "3. Out of stock tab — zero quantity items.\n"
            "4. Reorder or adjust stock as needed."
        ),
        "result": "Replenishment list before customers hit empty shelves.",
        "notes": "Tabs respect inventory module settings (alerts on/off).",
        "shot_file": "17-inventory-alerts.png",
    },
    # --- Customers & Invoicing ---
    {
        "sheet": "Finance & Reports",
        "section": "Customers",
        "module": "Customers",
        "route": "/customers",
        "task": "Add a credit customer",
        "role": "Manager, Sales",
        "steps": (
            "1. Customers → Add customer.\n"
            "2. Enter name, phone, email, address.\n"
            "3. Set credit limit if using invoicing.\n"
            "4. Save — attach at POS or Normal Sale checkout."
        ),
        "result": "Customer available for invoiced/credit sales.",
        "notes": "Outstanding balance shown on customer profile.",
        "shot_file": "18-customers.png",
    },
    {
        "sheet": "Finance & Reports",
        "section": "Invoices",
        "module": "Invoicing",
        "route": "/invoices",
        "task": "Create and send an invoice",
        "role": "Manager",
        "steps": (
            "1. Invoices → New invoice.\n"
            "2. Select customer, add line items (products).\n"
            "3. Save as draft.\n"
            "4. Send — posts receivable (Super Admin may need to approve if maker-checker applies).\n"
            "5. Record partial or full payments on the invoice."
        ),
        "result": "Customer balance updated; PDF available for download.",
        "notes": "Manager can create; invoice approve/send may require Super Admin.",
        "shot_file": "19-invoices.png",
    },
    {
        "sheet": "Finance & Reports",
        "section": "Expenses",
        "module": "Expenses",
        "route": "/expenses",
        "task": "Record an expense",
        "role": "Manager",
        "steps": (
            "1. Expenses → Add expense.\n"
            "2. Category, amount, date, description, payment method.\n"
            "3. Save — status Pending.\n"
            "4. Super Admin approves → affects reports."
        ),
        "result": "Expense logged; manager cannot approve own entries by default.",
        "notes": "Escalate pending expenses to Super Admin for approval.",
        "shot_file": "20-expenses.png",
    },
    {
        "sheet": "Finance & Reports",
        "section": "Income",
        "module": "Income",
        "route": "/income",
        "task": "Record other income",
        "role": "Manager",
        "steps": (
            "1. Income → Add income.\n"
            "2. Fill category, amount, date, notes.\n"
            "3. Submit for approval (Super Admin approves)."
        ),
        "result": "Non-POS income tracked in accounting reports.",
        "notes": "Same approval pattern as expenses.",
        "shot_file": "21-income.png",
    },
    {
        "sheet": "Finance & Reports",
        "section": "Accounting",
        "module": "Accounting",
        "route": "/accounting",
        "task": "View financial statements",
        "role": "Manager, Super Admin",
        "steps": (
            "1. Accounting → choose report tab.\n"
            "2. Balance Sheet, Income Statement, Trial Balance, Cash Flow.\n"
            "3. General Ledger / Account Statement for detail.\n"
            "4. Export PDF where available."
        ),
        "result": "Period financial position for management review.",
        "notes": "Requires accounting module and posted transactions.",
        "shot_file": "22-accounting.png",
    },
    {
        "sheet": "Finance & Reports",
        "section": "Reports",
        "module": "Reports",
        "route": "/reports",
        "task": "Run sales and inventory reports",
        "role": "Manager",
        "steps": (
            "1. Reports hub → pick report card.\n"
            "2. Set period: today, week, month.\n"
            "3. Sales summary, payment mix, daily sales, product performance, inventory overview.\n"
            "4. Legacy reports: profit/loss, tax, supplier, customer — via catalog link."
        ),
        "result": "Exportable operational reports for meetings.",
        "notes": "Each report can be toggled in Module Settings.",
        "shot_file": "23-reports-hub.png",
    },
    # --- Approvals ---
    {
        "sheet": "Approvals & Audit",
        "section": "Maker-checker",
        "module": "Approvals",
        "route": "/pending-approvals",
        "task": "Approve a price or stock change",
        "role": "Manager (checker), Super Admin",
        "steps": (
            "1. Pending approvals (under Reports or direct URL).\n"
            "2. Review queue: product price, stock adjustment, settings change, etc.\n"
            "3. Open item — see before/after, reason, who proposed.\n"
            "4. Approve or Reject (reason required on reject).\n"
            "5. Cannot approve your own change (except Super Admin bypass)."
        ),
        "result": "Approved change goes live; POS shows new values.",
        "notes": "Manager has products.approve and inventory.approve; not expenses/income.",
        "shot_file": "24-pending-approvals.png",
    },
    {
        "sheet": "Approvals & Audit",
        "section": "Audit log",
        "module": "Audit",
        "route": "/audit-log",
        "task": "Investigate who changed prices or stock",
        "role": "Manager, Super Admin",
        "steps": (
            "1. Audit log → filter by user, action, date.\n"
            "2. Review login, checkout, create/update, stock_adjust events.\n"
            "3. Use for dispute resolution and training."
        ),
        "result": "Traceability for sensitive operations.",
        "notes": "Manager cannot access User Management but can audit operations.",
        "shot_file": "25-audit-log.png",
    },
    # --- Super Admin ---
    {
        "sheet": "Super Admin",
        "section": "Users",
        "module": "Settings",
        "route": "/users",
        "task": "Create a new staff user",
        "role": "Super Admin",
        "steps": (
            "1. User Management → Add user.\n"
            "2. Username, name, email, password.\n"
            "3. Assign role: Manager, Sales Personnel, or custom role.\n"
            "4. Set active branch if multi-branch.\n"
            "5. Save — user can log in immediately."
        ),
        "result": "New staff member with correct permissions.",
        "notes": "Change default passwords in production.",
        "shot_file": "26-users.png",
    },
    {
        "sheet": "Super Admin",
        "section": "Roles",
        "module": "Settings",
        "route": "/roles",
        "task": "Customize role permissions",
        "role": "Super Admin",
        "steps": (
            "1. Role Management → select role or create custom role.\n"
            "2. Toggle permissions per module (view, create, update, approve, export).\n"
            "3. Save — may queue if Maker-checker protects role changes."
        ),
        "result": "Fine-grained access without code changes.",
        "notes": "Bootstrap roles: Super Admin, Manager, Sales Personnel.",
        "shot_file": "27-roles.png",
    },
    {
        "sheet": "Super Admin",
        "section": "Module settings",
        "module": "Settings",
        "route": "/module-settings",
        "task": "Enable or disable modules",
        "role": "Super Admin",
        "steps": (
            "1. Module Settings → toggle modules (Products, Sales, Inventory, etc.).\n"
            "2. Expand module → enable features (POS, variants, CSV, stock transfer).\n"
            "3. Apply preset (starter, retail, full) if desired.\n"
            "4. Changes refresh navigation for all users."
        ),
        "result": "Store only shows modules you need.",
        "notes": "Disabling a module hides routes and API access.",
        "shot_file": "28-module-settings.png",
    },
    {
        "sheet": "Super Admin",
        "section": "System settings",
        "module": "Settings",
        "route": "/system-settings",
        "task": "Configure store rules",
        "role": "Super Admin",
        "steps": (
            "1. System Settings → Store rules card.\n"
            "2. Maker-checker on/off, emergency stock mode.\n"
            "3. Allow sales to add products, payment methods at POS.\n"
            "4. Receipt: logo, header/footer, show SKU, auto-print.\n"
            "5. Per-module toggle cards below — save each section."
        ),
        "result": "Business rules enforced across POS and back office.",
        "notes": "Managers cannot open this page (route blocked).",
        "shot_file": "29-system-settings.png",
    },
    {
        "sheet": "Super Admin",
        "section": "Branches",
        "module": "Settings",
        "route": "/branches",
        "task": "Add a branch",
        "role": "Super Admin",
        "steps": (
            "1. Branch Management → Add branch.\n"
            "2. Code, name, address, manager.\n"
            "3. Mark headquarters if applicable.\n"
            "4. Staff switch branch from header selector."
        ),
        "result": "Multi-location inventory and sales separation.",
        "notes": "Requires multi_branch_support in Module Settings.",
        "shot_file": "30-branches.png",
    },
    {
        "sheet": "Super Admin",
        "section": "Financial approve",
        "module": "Finance",
        "route": "/expenses, /income, /invoices",
        "task": "Approve expenses, income, and invoices",
        "role": "Super Admin",
        "steps": (
            "1. Open Expenses or Income → filter Pending.\n"
            "2. Approve or reject each record.\n"
            "3. Invoices — Send/approve manual invoices.\n"
            "4. Money transfers — approve bank movements."
        ),
        "result": "Financial records post to accounting.",
        "notes": "Managers create; Super Admin approves by default.",
        "shot_file": "31-financial-approve.png",
    },
]


QUICK_REF = [
    ("Route", "Page", "Manager", "Super Admin", "Sales"),
    ("/", "Dashboard", "Yes", "Yes", "Yes (limited)"),
    ("/pos", "Retail POS", "Yes", "Yes", "Yes"),
    ("/pos/billing", "Terminal POS", "Yes", "Yes", "Yes"),
    ("/products", "Products", "Yes", "Yes", "If store allows"),
    ("/categories", "Categories", "Yes", "Yes", "If store allows"),
    ("/product-attributes", "Sizes & colors", "Yes", "Yes", "No"),
    ("/inventory", "Inventory", "Yes", "Yes", "No"),
    ("/barcodes", "Barcodes", "Yes", "Yes", "Limited"),
    ("/sales", "Sales history", "Yes", "Yes", "No"),
    ("/normal-sale", "Normal sale", "Yes", "Yes", "No"),
    ("/customers", "Customers", "Yes", "Yes", "Yes"),
    ("/invoices", "Invoices", "Yes", "Yes", "Limited"),
    ("/suppliers", "Suppliers", "Yes", "Yes", "No"),
    ("/employees", "Employees", "Yes", "Yes", "No"),
    ("/expenses", "Expenses", "Create only", "Full + approve", "No"),
    ("/income", "Income", "Create only", "Full + approve", "No"),
    ("/accounting", "Accounting", "Yes", "Yes", "No"),
    ("/reports", "Reports", "Yes", "Yes", "No"),
    ("/audit-log", "Audit log", "Yes", "Yes", "No"),
    ("/pending-approvals", "Pending approvals", "Yes", "Yes", "No"),
    ("/users", "Users", "No", "Yes", "No"),
    ("/roles", "Roles", "No", "Yes", "No"),
    ("/module-settings", "Module settings", "No", "Yes", "No"),
    ("/system-settings", "System settings", "No", "Yes", "No"),
    ("/branches", "Branches", "No", "Yes", "No"),
]


def style_header_row(ws, row: int, widths: dict) -> None:
    for col_idx, (_, header, width) in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER
        ws.column_dimensions[get_column_letter(col_idx)].width = width


def write_cover(wb: Workbook) -> None:
    ws = wb.active
    ws.title = "Cover"
    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 80
    rows = [
        ("Title", "CompleteByte POS — Manager & Super Admin User Manual"),
        ("Version", f"Generated {date.today().isoformat()}"),
        ("Product", "CompleteBytePOS retail / wholesale point of sale"),
        ("", ""),
        ("Purpose", "Onboarding guide for store managers and system administrators."),
        ("", ""),
        ("How to use this workbook", ""),
        ("", "• Read Getting Started, then your role sheet (Manager Daily Ops or Super Admin)."),
        ("", "• Each task row has step-by-step instructions and a screenshot column."),
        ("", "• Quick Reference lists every page and which roles can open it."),
        ("", ""),
        ("Default logins (change in production)", ""),
        ("Manager", "manager / manager123"),
        ("Super Admin", "admin / admin123"),
        ("Sales", "sales / sales123"),
        ("", ""),
        ("Refresh screenshots", ""),
        (
            "",
            "1. Start app: docker compose -f docker-compose.dev.yml up -d\n"
            "2. cd fe && node ../scripts/capture_user_manual_screenshots.js\n"
            "3. Re-run: python scripts/generate_user_manual.py",
        ),
        ("", ""),
        ("Support files", "docs/SETUP.md, docs/MAKER_CHECKER.md, docs/POS_UX_ROLES_AND_TESTING.md"),
    ]
    for r, (label, value) in enumerate(rows, start=1):
        ws.cell(row=r, column=1, value=label).font = SUBTITLE_FONT if label else Font()
        ws.cell(row=r, column=2, value=value).alignment = WRAP
        if label == "Title":
            ws.cell(row=r, column=1).font = TITLE_FONT
            ws.cell(row=r, column=2).font = TITLE_FONT


def write_data_sheet(wb: Workbook, sheet_name: str, rows: list[dict]) -> None:
    ws = wb.create_sheet(sheet_name)
    style_header_row(ws, 1, {})
    ws.freeze_panes = "A2"
    row_idx = 2
    for item in rows:
        for col_idx, (key, _, _) in enumerate(COLUMNS, start=1):
            val = item.get(key, "")
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.alignment = WRAP
            cell.border = BORDER
        ws.row_dimensions[row_idx].height = 90
        row_idx += 1


def embed_screenshots(wb: Workbook, shots_dir: Path) -> int:
    """Add Screenshot Gallery sheet and embed images into task sheets."""
    gallery = wb.create_sheet("Screenshot Gallery")
    gallery.append(["File", "Task", "Route", "Preview"])
    for col, w in enumerate([28, 36, 22, 80], start=1):
        gallery.column_dimensions[get_column_letter(col)].width = w
    g_row = 2
    embedded = 0

    shot_col = len(COLUMNS) + 1
    shot_header_col = len(COLUMNS)

    for sheet_name in wb.sheetnames:
        if sheet_name in ("Cover", "Quick Reference", "Screenshot Gallery"):
            continue
        ws = wb[sheet_name]
        if ws.max_row < 2:
            continue
        hdr = ws.cell(row=1, column=shot_header_col, value="Screenshot")
        hdr.font = HEADER_FONT
        hdr.fill = HEADER_FILL
        hdr.border = BORDER
        ws.column_dimensions[get_column_letter(shot_header_col)].width = 36

        for r in range(2, ws.max_row + 1):
            fname = ws.cell(row=r, column=len(COLUMNS)).value
            if not fname:
                continue
            path = shots_dir / str(fname)
            if not path.exists():
                make_placeholder(path, ws.cell(row=r, column=3).value or "", ws.cell(row=r, column=4).value or "")
            if not path.exists():
                continue
            try:
                img = XLImage(str(path))
                img.width = 280
                img.height = 158
                anchor = f"{get_column_letter(shot_col)}{r}"
                ws.add_image(img, anchor)
                ws.row_dimensions[r].height = 120
                task = ws.cell(row=r, column=4).value
                route = ws.cell(row=r, column=3).value
                gallery.cell(row=g_row, column=1, value=fname)
                gallery.cell(row=g_row, column=2, value=task)
                gallery.cell(row=g_row, column=3, value=route)
                gimg = XLImage(str(path))
                gimg.width = 480
                gimg.height = 270
                gallery.add_image(gimg, f"D{g_row}")
                gallery.row_dimensions[g_row].height = 205
                g_row += 1
                embedded += 1
            except Exception:
                pass
    return embedded


def write_quick_ref(wb: Workbook) -> None:
    ws = wb.create_sheet("Quick Reference")
    for r, row in enumerate(QUICK_REF, start=1):
        for c, val in enumerate(row, start=1):
            cell = ws.cell(row=r, column=c, value=val)
            cell.border = BORDER
            cell.alignment = WRAP
            if r == 1:
                cell.font = HEADER_FONT
                cell.fill = HEADER_FILL
    for col, w in zip(range(1, 6), [14, 24, 12, 14, 14]):
        ws.column_dimensions[get_column_letter(col)].width = w


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--screenshots-dir", type=Path, default=DEFAULT_SHOTS)
    args = parser.parse_args()
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    args.screenshots_dir.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    write_cover(wb)

    sheets = {}
    for row in MANUAL_ROWS:
        sheets.setdefault(row["sheet"], []).append(row)
    for name, rows in sheets.items():
        write_data_sheet(wb, name, rows)

    write_quick_ref(wb)
    count = embed_screenshots(wb, args.screenshots_dir)
    wb.save(OUT_FILE)
    print(f"Wrote {OUT_FILE}")
    print(f"Embedded {count} screenshots from {args.screenshots_dir}")


if __name__ == "__main__":
    main()
